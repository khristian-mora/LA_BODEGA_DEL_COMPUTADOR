"""
Script para scraping de información técnica e imágenes de productos
Uso: python scrape_products.py [--limit N] [--full]
"""
import openpyxl
import json
import re
import time
import sys
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

# Importar librerías de scraping
try:
    from duckduckgo_search import DDGS
    print("[OK] duckduckgo-search cargado")
except ImportError:
    print("[ERROR] Instalar: pip install duckduckgo-search")
    sys.exit(1)

try:
    import wikipediaapi
    print("[OK] wikipedia-api cargado")
except ImportError:
    print("[WARN] wikipedia-api no disponible (opcional)")

# ============ CONFIGURACIÓN ============
INPUT_FILE = 'LISTADO_CLEAN.xlsx'
OUTPUT_FILE = 'productos_enriquecidos.json'
OUTPUT_EXCEL = 'LISTADO_ENRIQUECIDO.xlsx'
MAX_RESULTS_IMAGES = 3  # Máximo imágenes por producto
MAX_WORKERS = 5  # Hilos paralelos para scraping

# ============ LISTA DE MARCAS PARA BÚSQUEDA ============
BRAND_SEARCH_MAP = {
    'HP': ['Hewlett-Packard', 'HP Inc'],
    'DELL': ['Dell Technologies'],
    'ASUS': ['ASUSTeK'],
    'ACER': ['Acer Inc'],
    'LENOVO': ['Lenovo Group'],
    'SAMSUNG': ['Samsung Electronics'],
    'LG': ['LG Electronics'],
    'SONY': ['Sony Corporation'],
    'CANON': ['Canon Inc'],
    'EPSON': ['Seiko Epson'],
    'BROTHER': ['Brother Industries'],
    'LOGITECH': ['Logitech International'],
    'KINGSTON': ['Kingston Technology'],
    'SANDISK': ['SanDisk Corporation'],
    'WD': ['Western Digital'],
    'SEAGATE': ['Seagate Technology'],
    'ADATA': ['ADATA Technology'],
    'TOSHIBA': ['Toshiba Corporation'],
    'INTEL': ['Intel Corporation'],
    'AMD': ['Advanced Micro Devices'],
    'NVIDIA': ['NVIDIA Corporation'],
    'MSI': ['Micro-Star International'],
    'GIGABYTE': ['Gigabyte Technology'],
    'CORSAIR': ['Corsair Gaming'],
    'RAZER': ['Razer Inc'],
    'GENIUS': ['Genius (brand)'],
    'TP-LINK': ['TP-Link Technologies'],
    'D-LINK': ['D-Link Corporation'],
    'TENDA': ['Tenda Technology'],
    'MERCUSYS': ['Mercusys'],
    'NOGANET': ['Noganet'],
    'JEWAY': ['Jeway'],
    'PANTRON': ['Pantron'],
    'COOLER MASTER': ['Cooler Master'],
    'NOCTUA': ['Noctua'],
    'NZXT': ['NZXT'],
    'PHILIPS': ['Philips'],
    'VIEWSONIC': ['ViewSonic'],
    'AOC': ['AOC International'],
    'BENQ': ['BenQ Corporation'],
    'APPLE': ['Apple Inc'],
    'HUAWEI': ['Huawei Technologies'],
    'XIAOMI': ['Xiaomi Corporation'],
    'NETGEAR': ['NETGEAR Inc'],
    'UBIQUITI': ['Ubiquiti Inc'],
    'CISCO': ['Cisco Systems'],
    'APC': ['APC (company)'],
    'THERMALTAKE': ['Thermaltake'],
    'FRACTAL DESIGN': ['Fractal Design'],
    'BE QUIET': ['be quiet!'],
    'SYNOLOGY': ['Synology Inc'],
    'QNAP': ['QNAP Systems'],
}

# ============ FUNCIÓN DE CLASIFICACIÓN ============
def classify_product(nombre):
    """Clasifica un producto por marca y categoría"""
    nombre_upper = nombre.upper().strip()
    
    # Buscar marca
    brand = 'GENERICO'
    brand_search_terms = []
    for b, terms in BRAND_SEARCH_MAP.items():
        if b in nombre_upper or any(t.upper() in nombre_upper for t in terms):
            brand = b
            brand_search_terms = terms
            break
    
    # Buscar categoría
    category = 'OTROS'
    subcategory = ''
    
    if any(kw in nombre_upper for kw in ['ADAPTADOR', 'CARGADOR', 'CHARGER']):
        category = 'ADAPTADORES/CARGADORES'
        if 'LAPTOP' in nombre_upper or 'NOTEBOOK' in nombre_upper:
            subcategory = 'Para Laptop'
        elif 'USB' in nombre_upper:
            subcategory = 'USB'
        elif '12V' in nombre_upper or '19V' in nombre_upper:
            subcategory = 'DC/AC'
    elif any(kw in nombre_upper for kw in ['MONITOR', 'LCD', 'LED']):
        category = 'MONITORES'
        if 'GAMER' in nombre_upper:
            subcategory = 'Gaming'
        elif 'IPS' in nombre_upper:
            subcategory = 'IPS'
        elif 'CURVO' in nombre_upper or 'CURVED' in nombre_upper:
            subcategory = 'Curvo'
    elif any(kw in nombre_upper for kw in ['TECLADO', 'KEYBOARD']):
        category = 'TECLADOS'
        if 'GAMER' in nombre_upper:
            subcategory = 'Gaming'
        elif 'MECANICO' in nombre_upper:
            subcategory = 'Mecánico'
    elif any(kw in nombre_upper for kw in ['MOUSE']):
        category = 'MOUSE'
        if 'GAMER' in nombre_upper:
            subcategory = 'Gaming'
        elif 'INALAMBRICO' in nombre_upper or 'WIRELESS' in nombre_upper:
            subcategory = 'Inalámbrico'
        elif 'OPTICO' in nombre_upper:
            subcategory = 'Óptico'
    elif any(kw in nombre_upper for kw in ['DISCO', 'HDD', 'SSD', 'DURO', 'STATE']):
        category = 'DISCOS DUROS'
        if 'SSD' in nombre_upper or 'STATE' in nombre_upper:
            subcategory = 'SSD'
        elif 'NVME' in nombre_upper:
            subcategory = 'NVMe'
        elif 'EXTERNO' in nombre_upper:
            subcategory = 'Externo'
    elif any(kw in nombre_upper for kw in ['MEMORIA', 'RAM', 'DDR']):
        category = 'MEMORIAS RAM'
        if 'DDR4' in nombre_upper:
            subcategory = 'DDR4'
        elif 'DDR3' in nombre_upper:
            subcategory = 'DDR3'
        elif 'DDR5' in nombre_upper:
            subcategory = 'DDR5'
    elif any(kw in nombre_upper for kw in ['IMPRESORA', 'PRINTER', 'FAX']):
        category = 'IMPRESORAS'
        if 'LASER' in nombre_upper:
            subcategory = 'Láser'
        elif 'INKJET' in nombre_upper or 'TINTA' in nombre_upper:
            subcategory = 'Tinta'
    elif any(kw in nombre_upper for kw in ['CARTUCHO', 'TONER']):
        category = 'CONSUMIBLES'
        subcategory = 'Tóner/Cartucho' if 'TONER' in nombre_upper else 'Cartucho'
    elif any(kw in nombre_upper for kw in ['CABLE', 'USB', 'HDMI', 'VGA', 'DVI', 'DISPLAYPORT', 'RJ45', 'PARALELO', 'SERIAL']):
        category = 'CABLES/CONECTORES'
        if 'HDMI' in nombre_upper:
            subcategory = 'HDMI'
        elif 'USB' in nombre_upper:
            subcategory = 'USB'
        elif 'VGA' in nombre_upper:
            subcategory = 'VGA'
        elif 'RJ45' in nombre_upper or 'ETHERNET' in nombre_upper:
            subcategory = 'Ethernet'
    elif any(kw in nombre_upper for kw in ['ROUTER', 'SWITCH', 'ACCESS POINT', 'WIFI', 'EXTENSOR', 'REPETIDOR']):
        category = 'REDES'
        if 'WIFI' in nombre_upper:
            subcategory = 'WiFi'
        elif 'SWITCH' in nombre_upper:
            subcategory = 'Switch'
        elif 'ROUTER' in nombre_upper:
            subcategory = 'Router'
    elif any(kw in nombre_upper for kw in ['WEBCAM', 'CAMARA']):
        category = 'CAMARAS/Webcam'
    elif any(kw in nombre_upper for kw in ['AURICULAR', 'HEADSET', 'MICROFONO']):
        category = 'AUDIO'
    elif any(kw in nombre_upper for kw in ['FUENTE', 'PSU', 'POWERSUPPLY']):
        category = 'FUENTES DE PODER'
    elif any(kw in nombre_upper for kw in ['TOWER', 'CASE', 'GABINETE']):
        category = 'GABINETES'
    elif any(kw in nombre_upper for kw in ['PROCESADOR', 'CPU']):
        category = 'PROCESADORES'
    elif any(kw in nombre_upper for kw in ['PLACA', 'MOTHERBOARD']):
        category = 'PLACAS BASE'
    elif any(kw in nombre_upper for kw in ['TARJETA DE VIDEO', 'GRAFICA', 'GPU', 'GEFORCE', 'RADEON']):
        category = 'TARJETAS DE VIDEO'
    elif any(kw in nombre_upper for kw in ['LAPTOP', 'NOTEBOOK', 'PORTATIL']):
        category = 'LAPTOPS'
    elif any(kw in nombre_upper for kw in ['TABLET']):
        category = 'TABLETS'
    elif any(kw in nombre_upper for kw in ['SILLA', 'ESCRITORIO', 'MUEBLE']):
        category = 'MUEBLES/OFICINA'
    elif any(kw in nombre_upper for kw in ['UPS', 'NOBREAK']):
        category = 'UPS/NOBREAK'
    elif any(kw in nombre_upper for kw in ['PENDRIVE', 'FLASH DRIVE']):
        category = 'MEMORIAS USB'
    elif any(kw in nombre_upper for kw in ['LECTOR', 'ESCANER', 'SCANNER', 'CODEBAR']):
        category = 'LECTORES/ESCANERS'
    elif any(kw in nombre_upper for kw in ['PROYECTOR']):
        category = 'PROYECTORES'
    elif any(kw in nombre_upper for kw in ['PARLANTE', 'SPEAKER']):
        category = 'PARLANTES'
    elif any(kw in nombre_upper for kw in ['GRABADORA', 'DVR', 'NVR', 'CCTV']):
        category = 'SEGURIDAD/VIGILANCIA'
    elif any(kw in nombre_upper for kw in ['BATERIA', 'PILA']):
        category = 'BATERIAS/PILAS'
    elif any(kw in nombre_upper for kw in ['CINTA']):
        category = 'CONSUMIBLES'
    
    return {
        'brand': brand,
        'brand_search': brand_search_terms,
        'category': category,
        'subcategory': subcategory
    }

# ============ SCRAPING DE INFORMACIÓN TÉCNICA ============
def get_wikipedia_info(product_name, brand_search_terms):
    """Obtiene información técnica de Wikipedia"""
    try:
        wiki_wiki = wikipediaapi.Wikipedia(
            language='es',
            extract_format=wikipediaapi.ExtractFormat.WIKI
        )
        
        # Intentar buscar por nombre del producto
        search_terms = [product_name]
        if brand_search_terms:
            search_terms.extend(brand_search_terms)
        
        for term in search_terms[:3]:  # Limitar a 3 intentos
            page = wiki_wiki.page(term)
            if page.exists():
                summary = page.summary[:500] if page.summary else ''
                return {
                    'title': page.title,
                    'summary': summary,
                    'url': page.fullurl
                }
        
        return None
    except Exception as e:
        return None

# ============ SCRAPING DE IMÁGENES ============
def search_images(product_name, brand, category, max_results=3):
    """Busca imágenes usando DuckDuckGo"""
    images = []
    try:
        # Construir query de búsqueda
        query = f"{brand} {product_name}"
        if brand == 'GENERICO':
            query = f"{product_name} {category}"
        
        # Limitar query a términos significativos
        query = ' '.join(query.split()[:8])
        
        with DDGS() as ddgs:
            results = list(ddgs.images(
                keywords=query,
                region='wt-wt',
                safesearch='off',
                max_results=max_results
            ))
            
            for r in results:
                if r.get('image'):
                    images.append({
                        'url': r['image'],
                        'thumbnail': r.get('thumbnail', ''),
                        'title': r.get('title', ''),
                        'source': r.get('source', '')
                    })
    except Exception as e:
        pass
    
    return images

# ============ PROCESAR UN PRODUCTO ============
def process_product(product, index, total):
    """Procesa un producto individual: clasifica y hace scraping"""
    nombre = product['nombre']
    precio = product['precio_num']
    
    # Clasificar
    classification = classify_product(nombre)
    
    # Información base
    result = {
        'index': index,
        'nombre': nombre,
        'precio': precio,
        'marca': classification['brand'],
        'categoria': classification['category'],
        'subcategoria': classification['subcategory'],
        'descripcion': '',
        'especificaciones': '',
        'url_info': '',
        'imagenes': []
    }
    
    # Solo hacer scraping para productos con marca conocida (no genéricos)
    if classification['brand'] != 'GENERICO' and index <= 500:  # Limitar a primeros 500
        # Buscar info en Wikipedia (opcional, más lento)
        # wiki_info = get_wikipedia_info(nombre, classification['brand_search'])
        # if wiki_info:
        #     result['descripcion'] = wiki_info['summary']
        #     result['url_info'] = wiki_info['url']
        
        # Buscar imágenes
        images = search_images(
            nombre, 
            classification['brand'], 
            classification['category'],
            max_results=MAX_RESULTS_IMAGES
        )
        result['imagenes'] = images
        
        # Pequeña pausa para no saturar
        time.sleep(0.2)
    
    if index % 50 == 0:
        print(f"  Procesado {index}/{total} productos...")
    
    return result

# ============ MAIN ============
def main():
    # Parse arguments
    limit = None
    full_mode = False
    
    for arg in sys.argv[1:]:
        if arg.startswith('--limit='):
            limit = int(arg.split('=')[1])
        elif arg == '--full':
            full_mode = True
    
    print(f"\n{'='*60}")
    print("SCRAPING DE PRODUCTOS - La Bodega del Computador")
    print(f"{'='*60}")
    
    # Leer Excel
    print(f"\n[1/4] Leyendo {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb.active
    
    products = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        nombre, precio, precio_num = row
        if nombre and nombre != 'NOMBRE':
            products.append({
                'nombre': str(nombre).strip(),
                'precio': precio,
                'precio_num': precio_num
            })
    wb.close()
    
    total = len(products)
    if limit:
        products = products[:limit]
        print(f"  Modo prueba: procesando {len(products)} de {total} productos")
    else:
        print(f"  Total productos: {total}")
    
    # Procesar productos
    print(f"\n[2/4] Procesando productos (clasificación + scraping)...")
    results = []
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {}
        for i, product in enumerate(products):
            future = executor.submit(process_product, product, i+1, len(products))
            futures[future] = i
            
        for future in as_completed(futures):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                print(f"  Error: {e}")
    
    # Ordenar por índice original
    results.sort(key=lambda x: x['index'])
    
    # Estadísticas
    print(f"\n[3/4] Generando estadísticas...")
    brands_count = {}
    categories_count = {}
    with_images = 0
    
    for r in results:
        brand = r['marca']
        cat = r['categoria']
        brands_count[brand] = brands_count.get(brand, 0) + 1
        categories_count[cat] = categories_count.get(cat, 0) + 1
        if r['imagenes']:
            with_images += 1
    
    print(f"  Marcas únicas: {len(brands_count)}")
    print(f"  Categorías únicas: {len(categories_count)}")
    print(f"  Productos con imágenes: {with_images}/{len(results)}")
    
    # Guardar JSON
    print(f"\n[4/4] Guardando resultados...")
    output = {
        'metadata': {
            'total_productos': len(results),
            'con_imagenes': with_images,
            'marcas': len(brands_count),
            'categorias': len(categories_count),
            'fecha': time.strftime('%Y-%m-%d %H:%M:%S')
        },
        'estadisticas': {
            'marcas': dict(sorted(brands_count.items(), key=lambda x: -x[1])),
            'categorias': dict(sorted(categories_count.items(), key=lambda x: -x[1]))
        },
        'productos': results
    }
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  JSON guardado: {OUTPUT_FILE}")
    
    # Guardar Excel enriquecido
    print(f"\n  Generando Excel enriquecido...")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Productos"
    
    # Headers
    headers = ['Nombre', 'Precio', 'Marca', 'Categoría', 'Subcategoría', 
               'Descripción', 'URL Info', 'Imágenes URLs']
    for col, header in enumerate(headers, 1):
        ws_out.cell(row=1, column=col, value=header)
    
    # Datos
    for row_idx, r in enumerate(results, 2):
        ws_out.cell(row=row_idx, column=1, value=r['nombre'])
        ws_out.cell(row=row_idx, column=2, value=r['precio'])
        ws_out.cell(row=row_idx, column=3, value=r['marca'])
        ws_out.cell(row=row_idx, column=4, value=r['categoria'])
        ws_out.cell(row=row_idx, column=5, value=r['subcategoria'])
        ws_out.cell(row=row_idx, column=6, value=r['descripcion'])
        ws_out.cell(row=row_idx, column=7, value=r['url_info'])
        # Comprimir URLs de imágenes
        img_urls = '\n'.join([img['url'] for img in r['imagenes'][:3]])
        ws_out.cell(row=row_idx, column=8, value=img_urls)
    
    wb_out.save(OUTPUT_EXCEL)
    print(f"  Excel guardado: {OUTPUT_EXCEL}")
    
    print(f"\n{'='*60}")
    print(f"COMPLETADO: {len(results)} productos procesados")
    print(f"{'='*60}")
    
    return output

if __name__ == '__main__':
    main()
