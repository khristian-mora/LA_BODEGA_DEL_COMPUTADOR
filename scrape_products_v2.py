"""
Script mejorado para scraping de productos
- Busca imágenes para TODOS los productos (marca + genéricos)
- Busca descripciones técnicas desde fuentes web
- Guarda imágenes en carpeta local + URLs
"""
import openpyxl
import json
import re
import time
import sys
import os
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote_plus

# Importar librerías de scraping
try:
    from duckduckgo_search import DDGS
except ImportError:
    print("[ERROR] pip install duckduckgo-search")
    sys.exit(1)

# ============ CONFIGURACIÓN ============
INPUT_FILE = 'LISTADO_CLEAN.xlsx'
OUTPUT_FILE = 'productos_enriquecidos_v2.json'
OUTPUT_EXCEL = 'LISTADO_ENRIQUECIDO_V2.xlsx'
IMAGES_DIR = 'public/images/products'
MAX_IMAGES = 3
MAX_WORKERS = 3

# ============ LISTA DE MARCAS ============
KNOWN_BRANDS = [
    'HP', 'DELL', 'ASUS', 'ACER', 'LENOVO', 'SAMSUNG', 'LG', 'SONY',
    'CANON', 'EPSON', 'BROTHER', 'LOGITECH', 'KINGSTON', 'SANDISK', 'WD',
    'SEAGATE', 'ADATA', 'TOSHIBA', 'INTEL', 'AMD', 'NVIDIA', 'MSI',
    'GIGABYTE', 'CORSAIR', 'RAZER', 'GENIUS', 'TP-LINK', 'D-LINK',
    'TENDA', 'MERCUSYS', 'NOGANET', 'JEWAY', 'PANTRON', 'COOLER MASTER',
    'NOCTUA', 'NZXT', 'PHILIPS', 'VIEWSONIC', 'AOC', 'BENQ', 'APPLE',
    'HUAWEI', 'XIAOMI', 'NETGEAR', 'UBIQUITI', 'CISCO', 'APC',
    'THERMALTAKE', 'FRACTAL', 'BE QUIET', 'SYNOLOGY', 'QNAP',
    'HITACHI', 'CRUCIAL', 'HYPERX', 'STEELSERIES', 'TURTLE BEACH',
    'PIONEER', 'CREATIVE', 'PLANTRONICS', 'JBL', 'BOSE',
    'VEYO', 'GENIUS', 'TRUST', 'MICROSOFT', 'DEPO'
]

# ============ CLASIFICACIÓN ============
def classify_product(nombre):
    """Clasifica producto por marca y categoría"""
    nombre_upper = nombre.upper().strip()
    
    # Buscar marca
    brand = 'GENERICO'
    for b in KNOWN_BRANDS:
        if b in nombre_upper:
            brand = b
            break
    
    # Buscar categoría
    category = 'OTROS'
    subcategory = ''
    
    cat_rules = [
        (['ADAPTADOR', 'CARGADOR', 'CHARGER', 'POWER SUPPLY'], 'ADAPTADORES/CARGADORES'),
        (['MONITOR', 'LCD', 'LED', 'SCREEN'], 'MONITORES'),
        (['TECLADO', 'KEYBOARD'], 'TECLADOS'),
        (['MOUSE'], 'MOUSE'),
        (['DISCO DURO', 'HDD', 'SSD', 'STATE DRIVE', 'DISK'], 'DISCOS DUROS'),
        (['MEMORIA RAM', 'DDR', 'DIMM', 'SO-DIMM'], 'MEMORIAS RAM'),
        (['IMPRESORA', 'PRINTER', 'FAX'], 'IMPRESORAS'),
        (['CARTUCHO', 'TONER', 'INK'], 'CONSUMIBLES/TINTAS'),
        (['CABLE USB', 'CABLE HDMI', 'CABLE VGA', 'CABLE DVI', 'CABLE ETHERNET', 'CABLE RJ', 'CABLE PARALELO'], 'CABLES'),
        (['USB'], 'CABLES/USB'),
        (['HDMI'], 'CABLES/HDMI'),
        (['VGA'], 'CABLES/VGA'),
        (['ROUTER', 'ACCESS POINT', 'EXTENSOR WIFI', 'REPETIDOR'], 'REDES/ROUTER'),
        (['SWITCH'], 'REDES/SWITCH'),
        (['WEBCAM', 'CAMARA DIGITAL'], 'CAMARAS'),
        (['AURICULAR', 'HEADSET', 'MICROFONO', 'AUDIFONO'], 'AUDIO'),
        (['PARLANTE', 'SPEAKER'], 'AUDIO/PARLANTES'),
        (['FUENTE DE PODER', 'PSU', 'POWERSUPPLY'], 'FUENTES DE PODER'),
        (['GABINETE', 'CASE', 'TOWER'], 'GABINETES'),
        (['PROCESADOR', 'CPU'], 'PROCESADORES'),
        (['PLACA BASE', 'MOTHERBOARD', 'MAINBOARD'], 'PLACAS BASE'),
        (['TARJETA DE VIDEO', 'TARJETA GRAFICA', 'GPU', 'GEFORCE', 'RADEON', 'RTX', 'GTX'], 'TARJETAS DE VIDEO'),
        (['LAPTOP', 'NOTEBOOK', 'PORTATIL'], 'LAPTOPS'),
        (['TABLET'], 'TABLETS'),
        (['SILLA', 'ESCRITORIO'], 'MUEBLES'),
        (['UPS', 'NOBREAK'], 'UPS'),
        (['PENDRIVE', 'FLASH DRIVE', 'MEMORIA USB'], 'MEMORIAS USB'),
        (['LECTOR CODIGO', 'ESCANER', 'SCANNER', 'LECTOR BARRAS'], 'ESCANERS'),
        (['PROYECTOR'], 'PROYECTORES'),
        (['DVR', 'NVR', 'CCTV', 'CAMARA SEGURIDAD'], 'SEGURIDAD'),
        (['BATERIA', 'PILA'], 'BATERIAS'),
        (['DISPENSADOR'], 'OTROS'),
    ]
    
    for keywords, cat in cat_rules:
        if any(kw in nombre_upper for kw in keywords):
            category = cat
            break
    
    return brand, category

# ============ BÚSQUEDA DE IMÁGENES ============
def search_product_images(nombre, brand, category):
    """Busca imágenes del producto"""
    images = []
    
    # Construir queries de búsqueda (probar varias)
    queries = []
    
    if brand != 'GENERICO':
        queries.append(f"{brand} {nombre}")
        queries.append(f"{brand} {category.split('/')[0]}")
    else:
        # Para genéricos, usar nombre limpio
        clean_name = nombre.replace('PARA', '').strip()
        queries.append(clean_name)
        queries.append(f"{category.split('/')[0]} {nombre.split()[0] if len(nombre.split()) > 0 else ''}")
    
    for query in queries[:2]:  # Máximo 2 queries por producto
        try:
            with DDGS() as ddgs:
                results = list(ddgs.images(
                    keywords=query,
                    region='wt-wt',
                    safesearch='off',
                    max_results=MAX_IMAGES
                ))
                
                for r in results:
                    if r.get('image'):
                        images.append({
                            'url': r['image'],
                            'thumbnail': r.get('thumbnail', ''),
                            'title': r.get('title', ''),
                            'source': r.get('source', '')
                        })
                
                if len(images) >= MAX_IMAGES:
                    break
                    
        except Exception as e:
            pass
        
        time.sleep(0.1)
    
    return images[:MAX_IMAGES]

# ============ DESCARGAR IMÁGENES LOCALES ============
def download_image(url, filename, folder):
    """Descarga imagen a carpeta local"""
    try:
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, filename)
        
        if os.path.exists(filepath):
            return filepath
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10, stream=True)
        
        if response.status_code == 200:
            content_type = response.headers.get('content-type', '')
            ext = '.jpg'
            if 'png' in content_type:
                ext = '.png'
            elif 'webp' in content_type:
                ext = '.webp'
            
            filepath = filepath.rsplit('.', 1)[0] + ext
            
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(8192):
                    f.write(chunk)
            
            return filepath
    except Exception as e:
        pass
    return None

# ============ BUSCAR DESCRIPCIÓN TÉCNICA ============
def search_product_description(nombre, brand):
    """Busca descripción técnica del producto"""
    # Construir query para búsqueda web
    query = f"{brand} {nombre} especificaciones ficha tecnica"
    if brand == 'GENERICO':
        query = f"{nombre} especificaciones"
    
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(
                keywords=query,
                region='wt-wt',
                max_results=3
            ))
            
            for r in results:
                body = r.get('body', '')
                if len(body) > 50:  # Descripción válida
                    return body[:500]  # Limitar a 500 caracteres
    except:
        pass
    
    return ''

# ============ PROCESAR UN PRODUCTO ============
def process_product(product, index, total, download_images=False):
    """Procesa un producto individual"""
    nombre = product['nombre']
    precio = product['precio_num']
    
    # Clasificar
    brand, category = classify_product(nombre)
    
    result = {
        'index': index,
        'nombre': nombre,
        'precio': precio,
        'marca': brand,
        'categoria': category,
        'descripcion': '',
        'imagenes': [],
        'imagenes_locales': []
    }
    
    # Buscar imágenes
    images = search_product_images(nombre, brand, category)
    result['imagenes'] = images
    
    # Descargar imágenes localmente si se solicita
    if download_images and images:
        for i, img in enumerate(images[:2]):
            safe_name = re.sub(r'[^\w\-_\.]', '_', nombre)[:50]
            filename = f"{index:04d}_{safe_name}_{i}.jpg"
            local_path = download_image(img['url'], filename, IMAGES_DIR)
            if local_path:
                result['imagenes_locales'].append(f"/images/products/{filename}")
    
    # Buscar descripción (solo para productos con marca, para no saturar)
    if brand != 'GENERICO' and index <= 500:
        desc = search_product_description(nombre, brand)
        if desc:
            result['descripcion'] = desc
    
    # Progreso
    if index % 25 == 0:
        print(f"  [{index}/{total}] Último: {nombre[:40]}... ({len(images)} imgs)")
    
    return result

# ============ MAIN ============
def main():
    # Args
    limit = None
    download = '--download' in sys.argv
    
    for arg in sys.argv[1:]:
        if arg.startswith('--limit='):
            limit = int(arg.split('=')[1])
    
    print(f"\n{'='*60}")
    print("SCRAPING V2 - La Bodega del Computador")
    print(f"{'='*60}")
    
    # Leer Excel
    print(f"\n[1/4] Leyendo {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb.active
    
    products = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        nombre, precio, precio_num = row
        if nombre and nombre != 'NOMBRE':
            products.append({
                'nombre': str(nombre).strip(),
                'precio': precio,
                'precio_num': precio_num
            })
    wb.close()
    
    total = len(products)
    if limit:
        products = products[:limit]
        print(f"  Modo prueba: {len(products)} de {total}")
    else:
        print(f"  Total: {total} productos")
    
    # Procesar
    print(f"\n[2/4] Procesando productos...")
    print(f"  Workers: {MAX_WORKERS}")
    print(f"  Descarga imágenes: {'Sí' if download else 'No'}")
    
    results = []
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {}
        for i, product in enumerate(products):
            future = executor.submit(
                process_product, 
                product, i+1, len(products), 
                download
            )
            futures[future] = i
        
        for future in as_completed(futures):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                print(f"  Error: {e}")
    
    results.sort(key=lambda x: x['index'])
    
    # Estadísticas
    print(f"\n[3/4] Estadísticas...")
    brands = {}
    cats = {}
    with_imgs = 0
    with_desc = 0
    
    for r in results:
        brands[r['marca']] = brands.get(r['marca'], 0) + 1
        cats[r['categoria']] = cats.get(r['categoria'], 0) + 1
        if r['imagenes']:
            with_imgs += 1
        if r['descripcion']:
            with_desc += 1
    
    print(f"  Productos: {len(results)}")
    print(f"  Con imágenes: {with_imgs} ({with_imgs*100//len(results)}%)")
    print(f"  Con descripción: {with_desc} ({with_desc*100//len(results)}%)")
    print(f"  Marcas únicas: {len(brands)}")
    print(f"  Categorías únicas: {len(cats)}")
    
    # Guardar JSON
    print(f"\n[4/4] Guardando resultados...")
    
    output = {
        'metadata': {
            'total': len(results),
            'con_imagenes': with_imgs,
            'con_descripcion': with_desc,
            'marcas': len(brands),
            'categorias': len(cats),
            'fecha': time.strftime('%Y-%m-%d %H:%M:%S')
        },
        'estadisticas': {
            'marcas': dict(sorted(brands.items(), key=lambda x: -x[1])),
            'categorias': dict(sorted(cats.items(), key=lambda x: -x[1]))
        },
        'productos': results
    }
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  JSON: {OUTPUT_FILE}")
    
    # Guardar Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Productos Enriquecidos"
    
    headers = ['Nombre', 'Precio', 'Marca', 'Categoría', 'Descripción', 'Imágenes URLs']
    for col, h in enumerate(headers, 1):
        ws_out.cell(row=1, column=col, value=h)
    
    for row_idx, r in enumerate(results, 2):
        ws_out.cell(row=row_idx, column=1, value=r['nombre'])
        ws_out.cell(row=row_idx, column=2, value=r['precio'])
        ws_out.cell(row=row_idx, column=3, value=r['marca'])
        ws_out.cell(row=row_idx, column=4, value=r['categoria'])
        ws_out.cell(row=row_idx, column=5, value=r['descripcion'])
        img_urls = '\n'.join([img['url'] for img in r['imagenes'][:3]])
        ws_out.cell(row=row_idx, column=6, value=img_urls)
    
    wb_out.save(OUTPUT_EXCEL)
    print(f"  Excel: {OUTPUT_EXCEL}")
    
    print(f"\n{'='*60}")
    print(f"¡COMPLETADO! {len(results)} productos procesados")
    print(f"{'='*60}")

if __name__ == '__main__':
    main()
