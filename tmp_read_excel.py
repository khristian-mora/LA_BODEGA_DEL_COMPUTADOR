import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('LISTADO_CLEAN.xlsx', read_only=True, data_only=True)
    ws = wb.active
    print(f"Hojas: {wb.sheetnames}")
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    
    count = 0
    for row in ws.iter_rows(values_only=True):
        if count < 20:
            print(f"Row {count+1}: {row}")
            count += 1
        else:
            break
    wb.close()
except Exception as e:
    print(f"Error: {type(e).__name__}: {e}")
